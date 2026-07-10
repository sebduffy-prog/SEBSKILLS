#!/usr/bin/env python3
"""Static auditor for Excel financial models. Flags the FAST/ICAEW-style
defects that break integrity: hardcodes buried in formulas, error cells,
inconsistent formulas across a row, external links, and un-guarded volatiles.

No network, no Excel needed. Reads the .xlsx directly with openpyxl.
Usage:  python3 audit_model.py path/to/model.xlsx [--json]
"""
import argparse
import json
import re
import sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Need openpyxl:  python3 -m pip install --user openpyxl")

# Numbers embedded in a formula (not row/col refs, not part of a range/func).
HARDCODE_RE = re.compile(r"(?<![A-Za-z0-9_$:.])(\d+\.?\d*)(?![A-Za-z0-9_(])")
ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
VOLATILES = ("NOW(", "TODAY(", "RAND(", "RANDBETWEEN(", "OFFSET(", "INDIRECT(")
# Digits that are harmless to see inside a formula.
BENIGN = {"0", "1", "100", "12", "365", "360", "4", "2", "1000"}


def norm_formula(formula, cell):
    """R1C1-ish signature: replace each ref with its offset from `cell` so two
    cells running the identical relative calc get the SAME signature."""
    from openpyxl.utils.cell import coordinate_to_tuple, column_index_from_string

    base_col, base_row = cell.column, cell.row

    def repl(m):
        col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), m.group(4)
        c = column_index_from_string(col)
        r = int(row)
        cofs = col if col_abs else f"c{c - base_col:+d}"
        rofs = row if row_abs else f"r{r - base_row:+d}"
        return f"[{cofs}{rofs}]"

    ref = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")
    return ref.sub(repl, formula.upper())


def audit(path):
    wb = load_workbook(path, data_only=False)
    wb_vals = load_workbook(path, data_only=True)
    findings = []

    for ws in wb.worksheets:
        vs = wb_vals[ws.title]
        row_sigs = defaultdict(list)  # row -> [(col, signature)]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                cached = vs[cell.coordinate].value
                coord = f"{ws.title}!{cell.coordinate}"

                if isinstance(cached, str) and cached in ERROR_LITERALS:
                    findings.append(("ERROR", "critical", coord,
                                     f"cell evaluates to {cached}"))
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                f = v

                if "[" in f and "]" in f:
                    findings.append(("EXTERNAL_LINK", "high", coord,
                                     "formula references another workbook"))
                for vol in VOLATILES:
                    if vol in f.upper():
                        findings.append(("VOLATILE", "low", coord,
                                         f"uses volatile {vol[:-1]}()"))
                nums = [n for n in HARDCODE_RE.findall(f) if n not in BENIGN]
                if nums:
                    findings.append(("HARDCODE", "high", coord,
                                     f"literal {', '.join(sorted(set(nums)))} inside formula"))
                if f.upper().count("IF(") >= 4:
                    findings.append(("DEEP_NEST", "medium", coord,
                                     "4+ nested IF — break into flags/steps"))
                row_sigs[cell.row].append((cell.column, norm_formula(f, cell)))

        # Inconsistent formula across a contiguous run in the same row.
        for r, cols in row_sigs.items():
            if len(cols) < 3:
                continue
            cols.sort()
            sigs = [s for _, s in cols]
            majority = max(set(sigs), key=sigs.count)
            if sigs.count(majority) < len(sigs):
                for c, s in cols:
                    if s != majority:
                        findings.append(("INCONSISTENT_ROW", "high",
                                         f"{ws.title}!{get_column_letter(c)}{r}",
                                         "formula differs from the rest of its row"))
    return findings


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    findings = audit(args.path)

    if args.json:
        print(json.dumps([{"code": c, "severity": s, "cell": cell, "detail": d}
                          for c, s, cell, d in findings], indent=2))
    else:
        order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
        for c, s, cell, d in sorted(findings, key=lambda x: order[x[1]]):
            print(f"[{s.upper():8}] {c:16} {cell:20} {d}")
        print(f"\n{len(findings)} finding(s).", file=sys.stderr)
    sys.exit(1 if any(s in ("critical", "high") for _, s, _, _ in findings) else 0)


if __name__ == "__main__":
    main()
