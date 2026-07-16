#!/usr/bin/env python3
"""Build a fully-linked DCF + LBO valuation model as an .xlsx — formulas only.

Every output cell is an Excel formula string that references input cells, so the
workbook recalculates live when a user edits an assumption. No hardcoded results.

Usage:
    python3 build_model.py [out.xlsx]

Requires openpyxl (pip install openpyxl). Python 3.8+.
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")          # inputs (hardcoded assumptions)
BLACK = Font(color="000000")         # formulas / calculations
GREEN = Font(color="008000")         # links to other sheets
BOLD = Font(bold=True)
HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(bold=True, color="FFFFFF")
PCT = "0.0%"
MULT = '0.0"x"'
NUM = "#,##0.0"
YEARS = 5  # explicit forecast horizon


def _hdr(ws, cell, text):
    ws[cell] = text
    ws[cell].fill = HDR
    ws[cell].font = HDRF


def build_assumptions(ws):
    _hdr(ws, "A1", "OPERATING & VALUATION ASSUMPTIONS")
    rows = [
        ("Revenue (Year 0, $m)", "Rev0", 1000.0),
        ("Revenue growth %", "g_rev", 0.08),
        ("EBIT margin %", "ebit_m", 0.20),
        ("Tax rate %", "tax", 0.25),
        ("D&A as % of revenue", "da_pct", 0.05),
        ("CapEx as % of revenue", "capex_pct", 0.06),
        ("Change in NWC as % of Δrevenue", "nwc_pct", 0.10),
        ("EBITDA margin % (=EBIT+D&A)", "ebitda_m", 0.25),
        ("Risk-free rate %", "rf", 0.04),
        ("Equity risk premium %", "erp", 0.055),
        ("Levered beta", "beta", 1.10),
        ("Pre-tax cost of debt %", "kd", 0.06),
        ("Target debt / total capital %", "wd", 0.30),
        ("Perpetuity growth rate %", "g_term", 0.025),
        ("Exit EV/EBITDA multiple", "exit_mult", 9.0),
        ("Net debt today ($m)", "net_debt", 400.0),
        ("Shares outstanding (m)", "shares", 100.0),
        ("LBO entry EV/EBITDA multiple", "entry_mult", 8.0),
        ("LBO entry leverage (Debt/EBITDA)", "entry_lev", 4.0),
        ("Transaction fees % of EV", "fees_pct", 0.02),
        ("Mandatory debt amort % of opening/yr", "amort_pct", 0.05),
        ("Holding period (years)", "hold", 5),
    ]
    ws["A3"] = "Driver"; ws["B3"] = "Name"; ws["C3"] = "Value"
    for c in ("A3", "B3", "C3"):
        ws[c].font = BOLD
    for i, (label, name, val) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, name).font = Font(italic=True, color="808080")
        c = ws.cell(r, 3, val)
        c.font = BLUE
        c.number_format = PCT if isinstance(val, float) and abs(val) < 1 else NUM
        ws.defined_names.add_named_range(name, ws, f"$C${r}") if False else None
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    return {name: f"Assumptions!$C${4+i}" for i, (_, name, _) in enumerate(rows)}


def build_dcf(ws, A):
    """A: dict name -> absolute ref on Assumptions sheet."""
    _hdr(ws, "A1", "UNLEVERED DCF (FCFF)")
    # Year columns B..(B+YEARS). Column B = Year 0 (actual), C.. = forecast.
    ws["A3"] = "Line item"
    ws["A3"].font = BOLD
    for y in range(0, YEARS + 1):
        col = get_column_letter(2 + y)
        c = ws[f"{col}3"]
        c.value = f"Year {y}"
        c.font = BOLD
        c.alignment = Alignment(horizontal="right")
    # helper to place a row
    def row(r, label):
        ws.cell(r, 1, label)
        return r
    r_rev, r_ebit, r_tax, r_da, r_capex, r_nwc, r_fcff = 4, 5, 6, 7, 8, 9, 10
    labels = {r_rev: "Revenue", r_ebit: "EBIT", r_tax: "Less: cash tax on EBIT",
              r_da: "Add: D&A", r_capex: "Less: CapEx", r_nwc: "Less: Δ NWC",
              r_fcff: "Free cash flow to firm (FCFF)"}
    for r, l in labels.items():
        ws.cell(r, 1, l)
    ws.cell(r_fcff, 1).font = BOLD
    for y in range(0, YEARS + 1):
        col = get_column_letter(2 + y)
        prev = get_column_letter(1 + y)
        if y == 0:
            ws[f"{col}{r_rev}"] = f"={A['Rev0']}"
        else:
            ws[f"{col}{r_rev}"] = f"={prev}{r_rev}*(1+{A['g_rev']})"
        ws[f"{col}{r_ebit}"] = f"={col}{r_rev}*{A['ebit_m']}"
        ws[f"{col}{r_tax}"] = f"=-{col}{r_ebit}*{A['tax']}"
        ws[f"{col}{r_da}"] = f"={col}{r_rev}*{A['da_pct']}"
        ws[f"{col}{r_capex}"] = f"=-{col}{r_rev}*{A['capex_pct']}"
        if y == 0:
            ws[f"{col}{r_nwc}"] = 0
        else:
            ws[f"{col}{r_nwc}"] = f"=-({col}{r_rev}-{prev}{r_rev})*{A['nwc_pct']}"
        # FCFF = EBIT*(1-tax) + D&A - CapEx - dNWC   (tax/capex/nwc already signed)
        ws[f"{col}{r_fcff}"] = (f"={col}{r_ebit}+{col}{r_tax}+{col}{r_da}"
                                f"+{col}{r_capex}+{col}{r_nwc}")
        for rr in (r_rev, r_ebit, r_tax, r_da, r_capex, r_nwc, r_fcff):
            ws[f"{col}{rr}"].number_format = NUM

    # WACC block
    ws["A13"] = "WACC BUILD"; ws["A13"].font = BOLD
    ws["A14"] = "Cost of equity (CAPM)"
    ws["B14"] = f"={A['rf']}+{A['beta']}*{A['erp']}"; ws["B14"].number_format = PCT
    ws["A15"] = "After-tax cost of debt"
    ws["B15"] = f"={A['kd']}*(1-{A['tax']})"; ws["B15"].number_format = PCT
    ws["A16"] = "WACC"
    ws["B16"] = f"=(1-{A['wd']})*B14+{A['wd']}*B15"
    ws["B16"].number_format = PCT; ws["B16"].font = BOLD

    # Discounting: forecast years 1..YEARS in columns C..(C+YEARS-1)
    ws["A18"] = "Discount period (t)"
    ws["A19"] = "Discount factor 1/(1+WACC)^t"
    ws["A20"] = "PV of FCFF"
    for y in range(1, YEARS + 1):
        col = get_column_letter(2 + y)
        ws[f"{col}18"] = y
        ws[f"{col}19"] = f"=1/(1+$B$16)^{col}18"
        ws[f"{col}19"].number_format = "0.000"
        ws[f"{col}20"] = f"={col}{r_fcff}*{col}19"
        ws[f"{col}20"].number_format = NUM

    last = get_column_letter(2 + YEARS)  # last forecast column
    # Terminal value — two methods
    ws["A22"] = "TERMINAL VALUE"; ws["A22"].font = BOLD
    ws["A23"] = "Gordon growth TV (undiscounted)"
    ws["B23"] = f"={last}{r_fcff}*(1+{A['g_term']})/($B$16-{A['g_term']})"
    ws["B23"].number_format = NUM
    ws["A24"] = "Exit-multiple TV (EBITDA x mult)"
    ws["B24"] = f"={last}{r_rev}*{A['ebitda_m']}*{A['exit_mult']}"
    ws["B24"].number_format = NUM
    ws["A25"] = "TV used (Gordon)"; ws["B25"] = "=B23"; ws["B25"].number_format = NUM
    ws["A26"] = "PV of terminal value"
    ws["B26"] = f"=B25*{last}19"; ws["B26"].number_format = NUM

    # Bridge to equity value
    ws["A28"] = "Sum of PV of FCFF"
    ws["B28"] = f"=SUM(C20:{last}20)"; ws["B28"].number_format = NUM
    ws["A29"] = "Enterprise value"; ws["B29"] = "=B28+B26"
    ws["B29"].number_format = NUM; ws["B29"].font = BOLD
    ws["A30"] = "Less: net debt"; ws["B30"] = f"=-{A['net_debt']}"
    ws["B30"].number_format = NUM
    ws["A31"] = "Equity value"; ws["B31"] = "=B29+B30"; ws["B31"].number_format = NUM
    ws["A32"] = "Value per share"
    ws["B32"] = f"=B31/{A['shares']}"; ws["B32"].number_format = '"$"#,##0.00'
    ws["B32"].font = BOLD
    ws.column_dimensions["A"].width = 34
    for y in range(0, YEARS + 1):
        ws.column_dimensions[get_column_letter(2 + y)].width = 11


def build_lbo(ws, A):
    _hdr(ws, "A1", "LBO — SOURCES & USES, DEBT SCHEDULE, RETURNS")
    # Entry: EBITDA0 links to DCF Year 0 EBIT + D&A
    ws["A3"] = "Entry EBITDA (Year 0)"
    ws["B3"] = "=DCF!B5+DCF!B7"; ws["B3"].number_format = NUM; ws["B3"].font = GREEN
    ws["A4"] = "Entry enterprise value"
    ws["B4"] = f"=B3*{A['entry_mult']}"; ws["B4"].number_format = NUM
    ws["A5"] = "Transaction fees"
    ws["B5"] = f"=B4*{A['fees_pct']}"; ws["B5"].number_format = NUM

    ws["A7"] = "SOURCES"; ws["A7"].font = BOLD
    ws["A8"] = "New debt raised"
    ws["B8"] = f"=B3*{A['entry_lev']}"; ws["B8"].number_format = NUM
    ws["A9"] = "Sponsor equity (plug)"
    ws["B9"] = "=B11-B8"; ws["B9"].number_format = NUM; ws["B9"].font = BOLD
    ws["A10"] = "USES"; ws["A10"].font = BOLD
    ws["A11"] = "Total uses (EV + fees)"
    ws["B11"] = "=B4+B5"; ws["B11"].number_format = NUM

    # Debt schedule across the holding period
    ws["A13"] = "DEBT SCHEDULE"; ws["A13"].font = BOLD
    ws["A14"] = "Year"
    r_open, r_ebitda, r_int, r_amort, r_close = 15, 16, 17, 18, 19
    ws.cell(r_open, 1, "Opening debt")
    ws.cell(r_ebitda, 1, "EBITDA")
    ws.cell(r_int, 1, "Cash interest")
    ws.cell(r_amort, 1, "Mandatory amort")
    ws.cell(r_close, 1, "Closing debt")
    for y in range(1, YEARS + 1):
        col = get_column_letter(1 + y)  # B for year 1
        prev = get_column_letter(y)      # A col header spacer for year1 open
        ws[f"{col}14"] = y; ws[f"{col}14"].font = BOLD
        if y == 1:
            ws[f"{col}{r_open}"] = "=B8"
        else:
            ws[f"{col}{r_open}"] = f"={get_column_letter(y)}{r_close}"
        # EBITDA grows with revenue growth off entry EBITDA
        ws[f"{col}{r_ebitda}"] = f"=$B$3*(1+{A['g_rev']})^{col}14"
        ws[f"{col}{r_int}"] = f"={col}{r_open}*{A['kd']}"
        ws[f"{col}{r_amort}"] = f"=MIN({col}{r_open},$B$8*{A['amort_pct']})"
        ws[f"{col}{r_close}"] = f"={col}{r_open}-{col}{r_amort}"
        for rr in (r_open, r_ebitda, r_int, r_amort, r_close):
            ws[f"{col}{rr}"].number_format = NUM

    lastc = get_column_letter(1 + YEARS)
    # Exit & returns
    ws["A21"] = "EXIT & RETURNS"; ws["A21"].font = BOLD
    ws["A22"] = "Exit EBITDA"
    ws["B22"] = f"={lastc}{r_ebitda}"; ws["B22"].number_format = NUM
    ws["A23"] = "Exit enterprise value"
    ws["B23"] = f"=B22*{A['exit_mult']}"; ws["B23"].number_format = NUM
    ws["A24"] = "Less: exit net debt"
    ws["B24"] = f"=-{lastc}{r_close}"; ws["B24"].number_format = NUM
    ws["A25"] = "Exit equity value"
    ws["B25"] = "=B23+B24"; ws["B25"].number_format = NUM; ws["B25"].font = BOLD
    ws["A26"] = "MOIC (exit equity / sponsor equity)"
    ws["B26"] = "=B25/B9"; ws["B26"].number_format = MULT; ws["B26"].font = BOLD
    ws["A27"] = "IRR (single-flow, %)"
    ws["B27"] = f"=(B25/B9)^(1/{A['hold']})-1"
    ws["B27"].number_format = PCT; ws["B27"].font = BOLD
    # IRR from an explicit cash-flow row (year 0 outflow, year N inflow)
    ws["A29"] = "Equity cash flows (for IRR)"
    ws["B29"] = "=-B9"; ws["B29"].number_format = NUM
    for y in range(1, YEARS):
        ws[f"{get_column_letter(1+y)}29"] = 0
    ws[f"{lastc}29"] = "=B25"; ws[f"{lastc}29"].number_format = NUM
    ws["A30"] = "IRR (Excel IRR on cash flows)"
    ws["B30"] = f"=IRR(B29:{lastc}29)"; ws["B30"].number_format = PCT
    ws.column_dimensions["A"].width = 34
    for y in range(0, YEARS + 1):
        ws.column_dimensions[get_column_letter(2 + y)].width = 11


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "valuation_model.xlsx"
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Assumptions"
    A = build_assumptions(ws_a)
    build_dcf(wb.create_sheet("DCF"), A)
    build_lbo(wb.create_sheet("LBO"), A)
    wb.save(out)
    print(f"wrote {out} — sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
