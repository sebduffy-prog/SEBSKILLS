#!/usr/bin/env python3
"""Build a fully-linked 3-statement model (IS/BS/CF) in one Excel sheet, FORMULAS ONLY.

Every forecast cell is an Excel formula referencing driver cells or other statement
cells -- no hard-coded outputs. The balance sheet balances by construction (see the
accounting identity in SKILL.md); row `Check` proves it. A revolver cash-sweep makes
interest circular, so iterative calc is enabled and a CIRC switch lets you break it.

macOS python3 (3.9) + openpyxl. Usage: python3 build_model.py [out.xlsx]
"""
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers

YEARS = 5
B = 2                       # column B = Year 0 (opening / actuals)
def cc(i): return get_column_letter(B + i)          # forecast col for year i (1..YEARS)
def prev(i): return get_column_letter(B + i - 1)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model"

# --- enable iterative calc so the interest<->revolver circularity resolves ---
calc = wb.calculation
calc.iterate = True
calc.iterateCount = 100
calc.iterateDelta = 1e-6
calc.fullCalcOnLoad = True

# ---------------- Drivers (edit these) ----------------
D = {                       # label -> (row, value)
    "CIRC (1=on,0=break)": 0.0,  # set to 1 in Excel AFTER first open to switch interest on
    "Revenue growth":  0.10,
    "COGS % rev":      0.60,
    "Opex % rev":      0.15,
    "Dep % open PPE":  0.10,
    "Capex % rev":     0.12,
    "Tax rate":        0.25,
    "Interest rate":   0.05,
    "DSO (days)":      45,
    "DIO (days)":      60,
    "DPO (days)":      30,
    "Dividend payout": 0.20,
    "Min cash":        20.0,
}
r = 2
drow = {}
ws["A1"] = "DRIVERS"; ws["A1"].font = Font(bold=True)
for k, v in D.items():
    ws[f"A{r}"] = k
    ws[f"B{r}"] = v
    drow[k] = f"$B${r}"      # absolute ref, e.g. $B$3
    r += 1

def d(k): return drow[k]

# ---------------- Row map ----------------
def block(title, names, start):
    ws[f"A{start}"] = title; ws[f"A{start}"].font = Font(bold=True)
    rows = {n: start + 1 + i for i, n in enumerate(names)}
    for n, rr in rows.items():
        ws[f"A{rr}"] = "  " + n
    return rows, start + 1 + len(names) + 1

istart = r + 1
IS, r = block("INCOME STATEMENT", [
    "Revenue", "COGS", "Gross profit", "Opex", "EBITDA",
    "D&A", "EBIT", "Interest exp", "EBT", "Tax", "Net income"], istart)

BS, r = block("BALANCE SHEET", [
    "Cash", "Accounts receivable", "Inventory", "Net PPE", "Total assets",
    "Accounts payable", "Debt (revolver)", "Equity", "Total L&E", "Check (A-L&E)"], r)

CF, r = block("CASH FLOW", [
    "Net income (cf)", "D&A (cf)", "Change AR", "Change inv", "Change AP", "CFO",
    "Capex", "CFI", "Revolver draw/(repay)", "Dividends", "CFF",
    "Net change in cash", "Opening cash", "Closing cash"], r)

# ---------------- Year 0 opening balances (must balance) ----------------
rev0 = 100.0
ppe0 = 100.0
cash0 = D["Min cash"]
ar0  = rev0 * D["DSO (days)"] / 365
inv0 = rev0 * D["COGS % rev"] * D["DIO (days)"] / 365
ap0  = rev0 * D["COGS % rev"] * D["DPO (days)"] / 365
debt0 = 40.0
assets0 = cash0 + ar0 + inv0 + ppe0
equity0 = assets0 - ap0 - debt0        # equity is the year-0 plug so BS balances
opens = {IS["Revenue"]: rev0, BS["Cash"]: cash0, BS["Accounts receivable"]: ar0,
         BS["Inventory"]: inv0, BS["Net PPE"]: ppe0, BS["Accounts payable"]: ap0,
         BS["Debt (revolver)"]: debt0, BS["Equity"]: equity0}
for rr, v in opens.items():
    ws[f"B{rr}"] = round(v, 4)
ws[f"B{BS['Total assets']}"] = f"=B{BS['Cash']}+B{BS['Accounts receivable']}+B{BS['Inventory']}+B{BS['Net PPE']}"
ws[f"B{BS['Total L&E']}"] = f"=B{BS['Accounts payable']}+B{BS['Debt (revolver)']}+B{BS['Equity']}"
ws[f"B{BS['Check (A-L&E)']}"] = f"=B{BS['Total assets']}-B{BS['Total L&E']}"

# ---------------- Forecast formulas (year 1..YEARS) ----------------
for i in range(1, YEARS + 1):
    c, p = cc(i), prev(i)
    # Income statement
    ws[f"{c}{IS['Revenue']}"]      = f"={p}{IS['Revenue']}*(1+{d('Revenue growth')})"
    ws[f"{c}{IS['COGS']}"]         = f"=-{c}{IS['Revenue']}*{d('COGS % rev')}"
    ws[f"{c}{IS['Gross profit']}"] = f"={c}{IS['Revenue']}+{c}{IS['COGS']}"
    ws[f"{c}{IS['Opex']}"]         = f"=-{c}{IS['Revenue']}*{d('Opex % rev')}"
    ws[f"{c}{IS['EBITDA']}"]       = f"={c}{IS['Gross profit']}+{c}{IS['Opex']}"
    ws[f"{c}{IS['D&A']}"]          = f"=-{p}{BS['Net PPE']}*{d('Dep % open PPE')}"
    ws[f"{c}{IS['EBIT']}"]         = f"={c}{IS['EBITDA']}+{c}{IS['D&A']}"
    # interest on AVERAGE debt -> references closing debt -> circular; CIRC switch breaks it
    avg_debt = f"(({p}{BS['Debt (revolver)']}+{c}{BS['Debt (revolver)']})/2)"
    ws[f"{c}{IS['Interest exp']}"] = f"=-{d('CIRC (1=on,0=break)')}*{avg_debt}*{d('Interest rate')}"
    ws[f"{c}{IS['EBT']}"]          = f"={c}{IS['EBIT']}+{c}{IS['Interest exp']}"
    ws[f"{c}{IS['Tax']}"]          = f"=-MAX({c}{IS['EBT']},0)*{d('Tax rate')}"
    ws[f"{c}{IS['Net income']}"]   = f"={c}{IS['EBT']}+{c}{IS['Tax']}"

    # Balance sheet (working capital from drivers; cash & debt from CF)
    ws[f"{c}{BS['Cash']}"]                = f"={c}{CF['Closing cash']}"
    ws[f"{c}{BS['Accounts receivable']}"] = f"={c}{IS['Revenue']}*{d('DSO (days)')}/365"
    ws[f"{c}{BS['Inventory']}"]           = f"=-{c}{IS['COGS']}*{d('DIO (days)')}/365"
    ws[f"{c}{BS['Net PPE']}"]             = f"={p}{BS['Net PPE']}+{c}{IS['D&A']}+{c}{IS['Revenue']}*{d('Capex % rev')}"
    ws[f"{c}{BS['Total assets']}"]        = f"=SUM({c}{BS['Cash']}:{c}{BS['Net PPE']})"
    ws[f"{c}{BS['Accounts payable']}"]    = f"=-{c}{IS['COGS']}*{d('DPO (days)')}/365"
    ws[f"{c}{BS['Debt (revolver)']}"]     = f"={p}{BS['Debt (revolver)']}+{c}{CF['Revolver draw/(repay)']}"
    ws[f"{c}{BS['Equity']}"]              = f"={p}{BS['Equity']}+{c}{IS['Net income']}+{c}{CF['Dividends']}"
    ws[f"{c}{BS['Total L&E']}"]           = f"={c}{BS['Accounts payable']}+{c}{BS['Debt (revolver)']}+{c}{BS['Equity']}"
    ws[f"{c}{BS['Check (A-L&E)']}"]       = f"={c}{BS['Total assets']}-{c}{BS['Total L&E']}"

    # Cash flow
    ws[f"{c}{CF['Net income (cf)']}"] = f"={c}{IS['Net income']}"
    ws[f"{c}{CF['D&A (cf)']}"]        = f"=-{c}{IS['D&A']}"
    ws[f"{c}{CF['Change AR']}"]       = f"=-({c}{BS['Accounts receivable']}-{p}{BS['Accounts receivable']})"
    ws[f"{c}{CF['Change inv']}"]      = f"=-({c}{BS['Inventory']}-{p}{BS['Inventory']})"
    ws[f"{c}{CF['Change AP']}"]       = f"={c}{BS['Accounts payable']}-{p}{BS['Accounts payable']}"
    ws[f"{c}{CF['CFO']}"]             = f"=SUM({c}{CF['Net income (cf)']}:{c}{CF['Change AP']})"
    ws[f"{c}{CF['Capex']}"]           = f"=-{c}{IS['Revenue']}*{d('Capex % rev')}"
    ws[f"{c}{CF['CFI']}"]             = f"={c}{CF['Capex']}"
    ws[f"{c}{CF['Dividends']}"]       = f"=-MAX({c}{IS['Net income']},0)*{d('Dividend payout')}"
    # cash available before revolver = opening cash + CFO + CFI + dividends
    avail = f"({p}{BS['Cash']}+{c}{CF['CFO']}+{c}{CF['CFI']}+{c}{CF['Dividends']})"
    # draw if below min cash; else repay down to zero revolver
    draw = f"MAX({d('Min cash')}-{avail},0)"
    repay = f"-MIN(MAX({avail}-{d('Min cash')},0),{p}{BS['Debt (revolver)']})"
    ws[f"{c}{CF['Revolver draw/(repay)']}"] = f"={draw}+{repay}"
    ws[f"{c}{CF['CFF']}"]              = f"={c}{CF['Revolver draw/(repay)']}+{c}{CF['Dividends']}"
    ws[f"{c}{CF['Net change in cash']}"] = f"={c}{CF['CFO']}+{c}{CF['CFI']}+{c}{CF['CFF']}"
    ws[f"{c}{CF['Opening cash']}"]    = f"={p}{BS['Cash']}"
    ws[f"{c}{CF['Closing cash']}"]    = f"={c}{CF['Opening cash']}+{c}{CF['Net change in cash']}"

# ---------------- Cosmetics ----------------
ws.column_dimensions["A"].width = 26
for i in range(0, YEARS + 1):
    ws.column_dimensions[get_column_letter(B + i)].width = 12
ws.freeze_panes = "B1"
for rr in list(IS.values()) + list(BS.values()) + list(CF.values()):
    for i in range(0, YEARS + 1):
        ws[f"{get_column_letter(B+i)}{rr}"].number_format = "#,##0.0;(#,##0.0)"

out = sys.argv[1] if len(sys.argv) > 1 else "three_statement_model.xlsx"
wb.save(out)
print(f"saved {out}")
