#!/usr/bin/env python3
"""Build a deterministic what-if workbook: scenario toggle + native Data Tables + tornado + break-even.

Everything is a live Excel formula -- no baked-in outputs. Highlights:
  * CHOOSE scenario selector drives the active driver column (SWITCH variant shown too).
  * A NATIVE two-variable Data Table (Excel's {=TABLE(row,col)} array) via openpyxl
    DataTableFormula -- r1 = row input cell, r2 = column input cell.
  * A native one-variable (column-oriented) Data Table.
  * A tornado table (each driver +/-20% around Base) with a horizontal bar chart.
  * Closed-form break-even units/revenue.

Data Tables only recompute inside Excel, so fullCalcOnLoad is forced on. openpyxl writes
the TABLE array on its master (top-left) cell only; Excel fills the body on first open.

macOS python3 (3.9) + openpyxl 3.1+. Usage: python3 build_scenario.py [out.xlsx]
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.worksheet.formula import DataTableFormula
from openpyxl.chart import BarChart, Reference

OUT = sys.argv[1] if len(sys.argv) > 1 else "scenario_sensitivity.xlsx"
PERTURB = 0.20                                   # tornado swing = +/-20% of Base

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WhatIf"
wb.calculation.fullCalcOnLoad = True             # Data Tables must recompute on open

HEAD = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")
BOLD = Font(bold=True)
MONEY = '#,##0'


def title(cell, text):
    ws[cell] = text
    ws[cell].font = HEAD
    ws[cell].fill = HFILL


# ------------------------------------------------------------------ scenario selector
ws["A1"] = "Scenario (1=Base 2=Bull 3=Bear)"
ws["A1"].font = BOLD
ws["B1"] = 1                                      # <- change this cell to flip scenarios
ws["B1"].fill = PatternFill("solid", fgColor="FFF2CC")

# ------------------------------------------------------------------ driver block
title("A3", "Driver"); title("B3", "Active"); title("C3", "Base"); title("D3", "Bull"); title("E3", "Bear")
drivers = [                                       # label, base, bull, bear
    ("Units",          1000, 1400,  700),
    ("Price",            50,   55,   45),
    ("Var cost / unit",  30,   28,   34),
    ("Fixed cost",    15000, 15000, 15000),
]
for i, (label, base, bull, bear) in enumerate(drivers):
    r = 4 + i
    ws[f"A{r}"] = label
    # ACTIVE value = pick the scenario column with CHOOSE (works in every Excel version)
    ws[f"B{r}"] = f"=CHOOSE($B$1,C{r},D{r},E{r})"
    ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = base, bull, bear
    for col in "BCDE":
        ws[f"{col}{r}"].number_format = MONEY
UNITS, PRICE, VC, FIXED = "B4", "B5", "B6", "B7"

# SWITCH alternative (Excel 2019 / 365 only) -- same result as B5 via CHOOSE
ws["A8"] = "Price via SWITCH (2019+)"
ws["B8"] = "=SWITCH($B$1,1,C5,2,D5,3,E5)"
ws["B8"].number_format = MONEY

# ------------------------------------------------------------------ outputs
title("A10", "Output"); title("B10", "Value")
outputs = [
    ("Revenue",          f"={UNITS}*{PRICE}"),
    ("Variable costs",   f"={UNITS}*{VC}"),
    ("Contribution",     f"=B11-B12"),
    ("Fixed costs",      f"={FIXED}"),
    ("Profit",           f"=B13-B15"),
    ("CM / unit",        f"={PRICE}-{VC}"),
    ("Break-even units", f"=IFERROR({FIXED}/({PRICE}-{VC}),NA())"),
    ("Break-even rev",   f"=B18*{PRICE}"),
]
for i, (label, formula) in enumerate(outputs):
    r = 11 + i
    ws[f"A{r}"] = label
    ws[f"B{r}"] = formula
    ws[f"B{r}"].number_format = MONEY
ws["A15"].font = ws["B15"].font = BOLD           # Profit row emphasised
PROFIT = "B15"

# ------------------------------------------------------------------ TWO-VARIABLE data table
# Corner holds the output formula; top row feeds Price (r1), left column feeds Units (r2).
title("A21", "2-VAR DATA TABLE: Profit by Price (top) x Units (left)")
ws["G22"] = f"={PROFIT}"                          # top-left corner = link to output
ws["G22"].number_format = MONEY
prices = [40, 45, 50, 55, 60]                     # across the top row -> row input cell
for j, p in enumerate(prices):
    c = ws.cell(row=22, column=8 + j, value=p)    # H22..L22
    c.font = BOLD
units = [600, 800, 1000, 1200, 1400, 1600]        # down the left column -> column input cell
for k, u in enumerate(units):
    c = ws.cell(row=23 + k, column=7, value=u)    # G23..G28
    c.font = BOLD
# master cell of the TABLE array = first body cell (H23); ref spans the whole body
ws["H23"] = DataTableFormula(ref="H23:L28", dt2D=True, r1=PRICE, r2=UNITS, ca=True)

# ------------------------------------------------------------------ ONE-VARIABLE data table
# Column-oriented: input values down a column, output formula one row up & one col right.
title("A31", "1-VAR DATA TABLE: Profit by Price")
ws["H31"] = f"={PROFIT}"                          # output formula sits above the body
ws["H31"].number_format = MONEY
for k, p in enumerate([40, 45, 50, 55, 60, 65]):
    ws.cell(row=32 + k, column=7, value=p).font = BOLD    # G32..G37 input prices
# 1-var, column oriented => dt2D False, dtr False, single input cell in r1 (Price)
ws["H32"] = DataTableFormula(ref="H32:H37", dt2D=False, dtr=False, r1=PRICE, ca=True)

# ------------------------------------------------------------------ TORNADO (Base +/- PERTURB)
# Closed-form profit with ONE Base driver perturbed, others held at Base (C4..C7).
title("A40", "TORNADO: profit swing, each driver +/-20% of Base")
title("A41", "Driver"); title("B41", "Low"); title("C41", "High"); title("D41", "Swing")
lo, hi = 1 - PERTURB, 1 + PERTURB
tornado = [                                        # label, low-profit, high-profit
    ("Units",     f"=(C4*{lo})*(C5-C6)-C7",  f"=(C4*{hi})*(C5-C6)-C7"),
    ("Price",     f"=C4*((C5*{lo})-C6)-C7",  f"=C4*((C5*{hi})-C6)-C7"),
    # cost UP lowers profit, so "Low" uses the high cost multiplier
    ("Var cost",  f"=C4*(C5-(C6*{hi}))-C7",  f"=C4*(C5-(C6*{lo}))-C7"),
    ("Fixed",     f"=C4*(C5-C6)-(C7*{hi})",  f"=C4*(C5-C6)-(C7*{lo})"),
]
for i, (label, low, high) in enumerate(tornado):
    r = 42 + i
    ws[f"A{r}"] = label
    ws[f"B{r}"] = low
    ws[f"C{r}"] = high
    ws[f"D{r}"] = f"=ABS(C{r}-B{r})"              # swing magnitude (sort desc for a true tornado)
    for col in "BCD":
        ws[f"{col}{r}"].number_format = MONEY

chart = BarChart()
chart.type = "bar"                                # horizontal bars = tornado look
chart.title = "Tornado (profit swing)"
chart.add_data(Reference(ws, min_col=4, min_row=41, max_row=45), titles_from_data=True)
chart.set_categories(Reference(ws, min_col=1, min_row=42, max_row=45))
chart.legend = None
ws.add_chart(chart, "F41")

ws.column_dimensions["A"].width = 34
for col in "BCDE":
    ws.column_dimensions[col].width = 12

wb.save(OUT)
print(f"wrote {OUT}")
